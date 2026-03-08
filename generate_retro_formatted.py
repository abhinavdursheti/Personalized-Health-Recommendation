from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

output_file = "Project_Sprint_Retrospective_Formatted.xlsx"

def generate_formatted_retro():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sprint Retrospective"

    # Define colors
    header_blue = "00B0F0"
    light_blue = "9BC2E6"
    
    # Define fonts
    title_font = Font(bold=True)
    header_font = Font(bold=True)
    guideline_font = Font(italic=True, size=10)
    normal_font = Font(size=11)
    
    # Define fills
    header_fill = PatternFill(start_color=header_blue, end_color=header_blue, fill_type="solid")
    guideline_fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")
    
    # Define borders
    thin_border_side = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Set column widths
    ws.column_dimensions['A'].width = 3
    for col in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 40
    ws.column_dimensions['F'].width = 12

    # Row 2: Sprint Retrospective Title
    ws.merge_cells('C2:E2') # As seen in the image, it spans C to E.
    title_cell = ws['C2']
    title_cell.value = "Sprint Retrospective"
    title_cell.font = title_font
    title_cell.fill = header_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply border to merged cells is tricky, but we can just apply to the range
    for col in ['C', 'D', 'E']:
        ws[f'{col}2'].border = thin_border
        ws[f'{col}2'].fill = header_fill # Ensure fill covers merged area

    # Row 3: Headers
    headers = ["What went well", "What went poorly", "What ideas do you have", "How should we take action"]
    for i, header in enumerate(headers):
        col_letter = chr(ord('B') + i)
        cell = ws[f'{col_letter}3']
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Row 4: Guidelines
    guidelines = [
        "This section highlights the successes and positive outcomes from the sprint. It helps the team recognize achievements and identify practices that should be continued.",
        "This section identifies the challenges, roadblocks, or failures encountered during the sprint. It helps pinpoint areas that need improvement or change.",
        "This section is for brainstorming new approaches, tools, or strategies to enhance the team's efficiency, productivity, or project outcomes.",
        "This section outlines specific steps or solutions to address the issues and implement the ideas discussed, ensuring continuous improvement in future sprints."
    ]
    for i, guide in enumerate(guidelines):
        col_letter = chr(ord('B') + i)
        cell = ws[f'{col_letter}4']
        cell.value = guide
        cell.font = guideline_font
        cell.fill = guideline_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = thin_border

    # Adjust row heights
    ws.row_dimensions[4].height = 60

    # The data representing the retrospective points for the Health Recommendation Project
    data = [
        [
            "Implementation of Random Forest for diet plans exceeded accuracy targets.",
            "Decision Tree model for exercise initially recommended too intense workouts for beginners.",
            "Include more nuanced categorization of 'Fitness Level' in the UI to feed better data to the Exercise model.",
            "Retune the Decision Tree classifier parameters to soften beginner exercise difficulty (Assignee: DS Team)."
        ],
        [
            "Smooth integration of the user authentication and health profiling models.",
            "Encountered minor TemplateSyntaxError rendering the correlation charts.",
            "Implement deeper automated UI testing to catch chart rendering errors immediately.",
            "Add automated end-to-end tests using the generated Functional Test Cases Excel (Assignee: QA Team)."
        ],
        [
            "The Advanced Analytics dashboard (Recovery & Stability) was built ahead of schedule.",
            "Delay in getting enough initial synthetic data to train the sleep tracking module.",
            "Crowdsource a larger baseline dataset for the Sleep model to avoid edge-case bugs.",
            "Generate 10,000 extra synthetic data rows for sleep scheduling limits (Assignee: Backend)."
        ],
        [
            "Database schema efficiently handled the health and habit metrics storage.",
            "Responsive design on the Habit Sensitivity page broke on smaller mobile screens.",
            "Adopt a mobile-first approach when developing future analytics views.",
            "Refactor CSS of the Analytics dashboard using flexbox grids (Assignee: UI/UX)."
        ]
    ]

    start_row = 5
    for row_idx, row_data in enumerate(data):
        current_row = start_row + row_idx
        for col_idx, cell_value in enumerate(row_data):
            col_letter = chr(ord('B') + col_idx)
            cell = ws[f'{col_letter}{current_row}']
            cell.value = cell_value
            cell.font = normal_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border
            
        # Optional: set a reasonable row height based on content
        ws.row_dimensions[current_row].height = 45

    wb.save('Project_Sprint_Retrospective_Formatted_Updated.xlsx')
    print(f"Successfully created 'Project_Sprint_Retrospective_Formatted_Updated.xlsx'.")

if __name__ == "__main__":
    generate_formatted_retro()
