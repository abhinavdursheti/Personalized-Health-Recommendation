import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

output_file = "Project_Sprint_Retrospective.xlsx"

# The data representing the retrospective points for the Health Recommendation Project
retrospective_data = {
    "What went well": [
        "Implementation of Random Forest for diet plans exceeded accuracy targets.",
        "Smooth integration of the user authentication and health profiling models.",
        "The Advanced Analytics dashboard (Recovery & Stability) was built ahead of schedule.",
        "Database schema efficiently handled the health and habit metrics storage."
    ],
    "What went poorly": [
        "Decision Tree model for exercise initially recommended too intense workouts for beginners.",
        "Encountered minor TemplateSyntaxError rendering the correlation charts.",
        "Delay in getting enough initial synthetic data to train the sleep tracking module.",
        "Responsive design on the Habit Sensitivity page broke on smaller mobile screens."
    ],
    "What ideas do we have to improve?": [
        "Include more nuanced categorization of 'Fitness Level' in the UI to feed better data to the Exercise model.",
        "Implement deeper automated UI testing to catch chart rendering errors immediately.",
        "Crowdsource a larger baseline dataset for the Sleep model to avoid edge-case bugs.",
        "Adopt a mobile-first approach when developing future analytics views."
    ],
    "Action items": [
        "Retune the Decision Tree classifier parameters to soften beginner exercise difficulty (Assignee: DS Team).",
        "Add automated end-to-end tests using the generated Functional Test Cases Excel (Assignee: QA Team).",
        "Generate 10,000 extra synthetic data rows for sleep scheduling limits (Assignee: Backend).",
        "Refactor CSS of the Analytics dashboard using flexbox grids (Assignee: UI/UX)."
    ]
}

# Create a DataFrame
df = pd.DataFrame(retrospective_data)

# Write to Excel
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Sprint Retrospective')
    
    # Auto-adjust column widths for better readability and enable text wrapping
    worksheet = writer.sheets['Sprint Retrospective']
    
    # Set header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="00A2D9", end_color="00A2D9", fill_type="solid")
    
    for cell in worksheet["1:1"]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Adjust widths and row alignments
    for idx, col in enumerate(df.columns, 1):
        column_letter = get_column_letter(idx)
        worksheet.column_dimensions[column_letter].width = 45 # Make columns wide
        
        # Enable text wrapping for all cells in the column
        for cell in worksheet[column_letter]:
            if cell.row != 1:  # Not headers
                cell.alignment = Alignment(wrap_text=True, vertical="top")

print(f"Successfully created '{output_file}'.")
