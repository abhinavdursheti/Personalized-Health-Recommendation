from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

output_file = "Functional_Test_Cases_Grouped_45.xlsx"

def generate_grouped_test_cases():
    wb = Workbook()
    ws = wb.active
    ws.title = "Functional Test Cases"

    # Define colors
    group_header_color = "5A9E80"  # Darkish green
    col_header_color = "75B899"    # Lighter green
    
    # Define fonts
    group_header_font = Font(bold=True, color="FFFFFF", size=12)
    col_header_font = Font(bold=True, color="FFFFFF")
    normal_font = Font(size=11)
    
    # Define fills
    group_header_fill = PatternFill(start_color=group_header_color, end_color=group_header_color, fill_type="solid")
    col_header_fill = PatternFill(start_color=col_header_color, end_color=col_header_color, fill_type="solid")
    
    # Define borders
    thin_border_side = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Set column widths
    widths = [15, 25, 35, 35, 35, 10, 35]
    for i, w in enumerate(widths):
        col_letter = get_column_letter(i + 1)
        ws.column_dimensions[col_letter].width = w

    # Define exact 45 test cases logically grouped
    grouped_tests = {
        "1. User Registration & Login": [
            ("User Registration", "Navigate to Registration Page", "1. Open application.\n2. Click 'Register'.", "Registration page loads successfully.", "Registration page loads.", "Pass", "Check UI rendering."),
            ("User Registration", "Valid User Registration", "1. Enter valid details.\n2. Click 'Register'.", "Account created successfully.", "Account created.", "Pass", "Check DB."),
            ("User Registration", "Invalid Email Format", "1. Enter email without '@'.\n2. Click 'Register'.", "Validation error displayed.", "Error displayed.", "Pass", "Frontend validation."),
            ("User Registration", "Weak Password Rejection", "1. Enter short password.\n2. Click 'Register'.", "Password strength error shown.", "Error shown.", "Pass", "Backend validation."),
            ("User Registration", "Password Mismatch", "1. Enter non-matching passwords.\n2. Click 'Register'.", "Mismatch error displayed.", "Error displayed.", "Pass", "Frontend validation."),
            ("User Login", "Navigate to Login Page", "1. Open application.\n2. Click 'Login'.", "Login page loads successfully.", "Login page loads.", "Pass", "Check UI rendering."),
            ("User Login", "Valid User Login", "1. Enter valid credentials.\n2. Click 'Login'.", "Redirected to dashboard.", "Redirected to dashboard.", "Pass", "Auth token set."),
            ("User Login", "Invalid Credentials", "1. Enter wrong password.\n2. Click 'Login'.", "Invalid credentials error shown.", "Error shown.", "Pass", "Auth rejection.")
        ],
        "2. Health Profile Setup": [
            ("Health Profile", "Navigate to Profile Setup", "1. Click 'Profile' in menu.", "Profile setup form loads.", "Form loads.", "Pass", "UI Check."),
            ("Health Profile", "Enter Valid Age", "1. Input age in valid range.", "Age field accepts input.", "Input accepted.", "Pass", "Validation check."),
            ("Health Profile", "Enter Valid Weight", "1. Input weight.", "Weight field accepts input.", "Input accepted.", "Pass", "Validation check."),
            ("Health Profile", "Enter Valid Height", "1. Input height.", "Height field accepts input.", "Input accepted.", "Pass", "Validation check."),
            ("Health Profile", "Auto-calculate BMI", "1. Enter height and weight.", "BMI is automatically calculated.", "BMI calculated correctly.", "Pass", "JS calculation check."),
            ("Health Profile", "Select Activity Level", "1. Choose from dropdown.", "Dropdown selects properly.", "Selection saved.", "Pass", "UI Check."),
            ("Health Profile", "Save Profile", "1. Click 'Save Profile'.", "Profile saved to database.", "Profile saved.", "Pass", "DB verification.")
        ],
        "3. ML Diet Recommendations": [
            ("Diet Recommendation", "Trigger Diet Model", "1. Navigate to recommendations.\n2. Click 'Generate Diet'.", "Model starts forming predictions.", "Model triggered.", "Pass", "RF Model init."),
            ("Diet Recommendation", "View BMR Output", "1. Observe BMR reading.", "BMR is displayed based on profile.", "BMR displayed accurately.", "Pass", "Math Verification."),
            ("Diet Recommendation", "View TDEE Output", "1. Observe TDEE reading.", "TDEE is displayed correctly.", "TDEE displayed.", "Pass", "Math Verification."),
            ("Diet Recommendation", "View Calorie Intake", "1. Observe optimal calories.", "Calorie prediction shown.", "Prediction shown.", "Pass", "RF Model output."),
            ("Diet Recommendation", "View Macronutrients", "1. Observe protein/carb/fat split.", "Macros distributed based on goal.", "Macros displayed.", "Pass", "Algorithmic split.")
        ],
        "4. ML Exercise Recommendations": [
            ("Exercise Recommendation", "Trigger Exercise Model", "1. Click 'Generate Exercise'.", "Decision tree model triggers.", "Model triggered.", "Pass", "DT Model init."),
            ("Exercise Recommendation", "View Workout Type", "1. Observe workout suggestions.", "Appropriate workouts listed.", "Workouts listed.", "Pass", "Check against fitness level."),
            ("Exercise Recommendation", "View Duration", "1. Observe duration.", "Duration matches user time.", "Duration matches.", "Pass", "DT logic check."),
            ("Exercise Recommendation", "View Frequency", "1. Observe weekly frequency.", "Frequency matches fitness goals.", "Frequency matches.", "Pass", "DT logic check.")
        ],
        "5. ML Sleep Recommendations": [
            ("Sleep Recommendation", "Trigger Sleep Model", "1. Click 'Generate Sleep'.", "Linear Regression model triggers.", "Model triggered.", "Pass", "LR Model init."),
            ("Sleep Recommendation", "View Predicted Duration", "1. Observe sleep hours.", "Optimal hours predicted.", "Hours predicted.", "Pass", "LR Model output."),
            ("Sleep Recommendation", "View Suggested Bedtime", "1. Observe bedtime.", "Bedtime calculated based on cycles.", "Bedtime shown.", "Pass", "Logic check."),
            ("Sleep Recommendation", "View Wake-up Time", "1. Observe wake-up time.", "Wake-up aligns with bedtime.", "Wake-up shown.", "Pass", "Logic check.")
        ],
        "6. Daily Health Tracking": [
            ("Health Tracking", "Navigate to Daily Log", "1. Click 'Daily Log'.", "Log entry form loads.", "Form loads.", "Pass", "UI Check."),
            ("Health Tracking", "Enter Daily Weight", "1. Subject weight input.", "Input accepted.", "Input accepted.", "Pass", "Validation."),
            ("Health Tracking", "Enter Sleep Hours", "1. Subject sleep input.", "Input accepted.", "Input accepted.", "Pass", "Validation."),
            ("Health Tracking", "Enter Exercise Min", "1. Subject exercise input.", "Input accepted.", "Input accepted.", "Pass", "Validation."),
            ("Health Tracking", "Enter Calorie Intake", "1. Subject calorie input.", "Input accepted.", "Input accepted.", "Pass", "Validation."),
            ("Health Tracking", "Save Daily Log", "1. Click 'Submit'.", "Data saved successfully.", "Data saved.", "Pass", "DB check.")
        ],
        "7. Analytics Dashboard": [
            ("Analytics Dashboard", "View Dashboard", "1. Navigate to dashboard.", "Dashboard loads.", "Dashboard loads.", "Pass", "UI Check."),
            ("Analytics Dashboard", "View Progress Charts", "1. Open dashboard window.", "Charts render with historical data.", "Charts render.", "Pass", "Chart.js check."),
            ("Analytics Dashboard", "View Recent Logs", "1. Scroll to logs table.", "Table shows recent day entries.", "Table populated.", "Pass", "Query check.")
        ],
        "8. Advanced Analytics - Recovery Engine": [
            ("Recovery Engine", "Trigger Analysis", "1. Click 'Recovery Analysis'.", "Engine processes history data.", "Engine processed.", "Pass", "RF Model init."),
            ("Recovery Engine", "View Recovery Days", "1. Observe recovery days result.", "Days predicted properly.", "Days shown.", "Pass", "RF Model output."),
            ("Recovery Engine", "View Stability Score", "1. Observe stability percentage.", "Score calculates 0-100%.", "Score shown.", "Pass", "GB Classifier check."),
            ("Recovery Engine", "View Consistency Rate", "1. Observe logging consistency.", "Rate calculates correctly.", "Rate shown.", "Pass", "Metric calculation.")
        ],
        "9. Advanced Analytics - Correlation & Habit": [
            ("Correlation Engine", "Trigger Correlation Analysis", "1. Click 'Correlation Engine'.", "SciPy pearson correlation runs.", "Analysis runs.", "Pass", "Script execution."),
            ("Correlation Engine", "View Top Correlations", "1. Observe behavior matches.", "Highest correlated items split.", "Correlations shown.", "Pass", "Data validity."),
            ("Habit Sensitivity", "Trigger Habit Analysis", "1. Click 'Habit Analyzer'.", "Fragility RF Model triggers.", "Model triggers.", "Pass", "RF Model init."),
            ("Habit Sensitivity", "View Fragility Ranking", "1. Observe habit list.", "Habits sorted resilient to fragile.", "Habits sorted.", "Pass", "Output processing.")
        ]
    }

    current_row = 1
    col_headers = ["Feature", "Test Case", "Steps to Execute", "Expected Output", "Actual Output", "Status", "More Information"]

    for group_name, tests in grouped_tests.items():
        # 1. Add Group Header Row (Spans A-G)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        group_cell = ws.cell(row=current_row, column=1)
        group_cell.value = group_name
        group_cell.font = group_header_font
        group_cell.fill = group_header_fill
        group_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, 8):
            ws.cell(row=current_row, column=col).border = thin_border
            
        current_row += 1

        # 2. Add Column Headers Row
        for col_idx, header in enumerate(col_headers, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.font = col_header_font
            cell.fill = col_header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        current_row += 1

        # 3. Add Data Rows
        for test_data in tests:
            for col_idx, cell_value in enumerate(test_data, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = cell_value
                cell.font = normal_font
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = thin_border
            ws.row_dimensions[current_row].height = 40
            current_row += 1
            
        # Add a blank separation row (optional, matching image separation)
        current_row += 1 

    wb.save(output_file)
    print(f"Successfully created '{output_file}'.")

if __name__ == "__main__":
    generate_grouped_test_cases()
