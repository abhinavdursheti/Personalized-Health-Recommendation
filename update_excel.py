import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

# Define the columns based on the provided template
columns = [
    "Feature",
    "Test Case",
    "Steps to execute test case",
    "Expected Output",
    "Actual Output",
    "Status",
    "More Information"
]

# Define the test cases based on project features
test_cases = [
    {
        "Feature": "User Management",
        "Test Case": "Valid User Registration",
        "Steps to execute test case": "1. Open the application's registration page.\n2. Enter valid user details (username, email, password, confirm password).\n3. Click on 'Register' button.",
        "Expected Output": "A new user account should be created.\n\nThe user should be redirected to the login page or automatically logged in.",
        "Actual Output": "A new user account is successfully created.\n\nThe user is redirected to the login page.",
        "Status": "Pass",
        "More Information": "Check the database to ensure the user is saved securely."
    },
    {
        "Feature": "User Management",
        "Test Case": "Valid User Login",
        "Steps to execute test case": "1. Open the application's login page.\n2. Enter a valid username.\n3. Enter a valid password.\n4. Click on the 'Login' button.",
        "Expected Output": "The user should be successfully logged into the system.\n\nThe application should redirect the user to the dashboard view.",
        "Actual Output": "The user is successfully logged in.\n\nThe application redirects the user to the dashboard.",
        "Status": "Pass",
        "More Information": "No error messages are displayed.\n\nThe user dashboard information displays correctly."
    },
    {
        "Feature": "Health Profile",
        "Test Case": "Setup Health Profile",
        "Steps to execute test case": "1. Navigate to health profile setup.\n2. Enter age, weight, height, activity level, and health goals.\n3. Save profile.",
        "Expected Output": "Health profile is successfully saved.\n\nBMI is successfully calculated from height and weight.",
        "Actual Output": "Health profile is successfully saved.\n\nBMI displays correctly based on input.",
        "Status": "Pass",
        "More Information": "Verify the recorded values match in the database profile."
    },
    {
        "Feature": "ML Recommendations",
        "Test Case": "Generate Diet Plan",
        "Steps to execute test case": "1. Navigate to recommendations section.\n2. Trigger Diet Recommendation model.\n3. Review generated meal plans.",
        "Expected Output": "Model calculates BMR and TDEE.\n\nPersonalized macronutrient breakdown and daily calorie intake are shown.",
        "Actual Output": "BMR and TDEE are calculated.\n\nMacronutrient distribution and optimal calorie intake are correctly displayed.",
        "Status": "Pass",
        "More Information": "Validates Random Forest Regressor execution based on user profile."
    },
    {
        "Feature": "ML Recommendations",
        "Test Case": "Generate Exercise Plan",
        "Steps to execute test case": "1. Navigate to recommendations section.\n2. Trigger Exercise Recommendation model.\n3. Review workout plans.",
        "Expected Output": "Customized workout plan is presented based on fitness level and goals.\n\nExercise frequency is recommended.",
        "Actual Output": "Customized workout plan is presented.\n\nExercise frequency suggestions match user goals.",
        "Status": "Pass",
        "More Information": "Validates Decision Tree Classifier output."
    },
    {
        "Feature": "ML Recommendations",
        "Test Case": "Generate Sleep Schedule",
        "Steps to execute test case": "1. Navigate to recommendations section.\n2. Trigger Sleep Recommendation model.\n3. Review sleep schedule.",
        "Expected Output": "Optimal sleep duration is predicted.\n\nBedtime and wake-up times are suggested.",
        "Actual Output": "Optimal sleep duration is predicted.\n\nBedtime and wake-up schedule is generated.",
        "Status": "Pass",
        "More Information": "Validates Linear Regression model execution."
    },
    {
        "Feature": "Health Tracking",
        "Test Case": "Log Daily Metrics",
        "Steps to execute test case": "1. Navigate to dashboard or tracking entry.\n2. Enter weight, sleep hours, exercise minutes, and calories.\n3. Submit data.",
        "Expected Output": "Daily health log is saved successfully.\n\nProgress charts update based on new data.",
        "Actual Output": "Daily health log is saved.\n\nDashboard progress charts refresh with new data point.",
        "Status": "Pass",
        "More Information": "Ensure real-time health metrics display is functional."
    },
    {
        "Feature": "Advanced Analytics ML",
        "Test Case": "Recovery & Stability Engine",
        "Steps to execute test case": "1. Navigate to Analytics section.\n2. Click on Recovery & Stability Prediction.\n3. Generate Analysis.",
        "Expected Output": "Recovery days needed and behavior stability score (0-100%) are calculated and displayed.\n\nConsistency and adherence rates shown.",
        "Actual Output": "Recovery days needed and behavior stability score are calculated and displayed.\n\nConsistency and adherence rates shown without errors.",
        "Status": "Pass",
        "More Information": "Checks Random Forest Regressor and Gradient Boosting Classifier."
    },
    {
        "Feature": "Advanced Analytics ML",
        "Test Case": "Correlation Engine",
        "Steps to execute test case": "1. Navigate to Analytics section.\n2. Click on Behavior-Cause Correlation Engine.\n3. Generate Analysis.",
        "Expected Output": "Correlation coefficients between behaviors and outcomes are identified.\n\nRoot causes are displayed.",
        "Actual Output": "Correlation coefficients and root causes are identified successfully.",
        "Status": "Pass",
        "More Information": "Actionable recommendations should be generated based on stats."
    },
    {
        "Feature": "Advanced Analytics ML",
        "Test Case": "Habit Sensitivity Analyzer",
        "Steps to execute test case": "1. Navigate to Analytics section.\n2. Click on Habit Sensitivity Analyzer.\n3. Generate Analysis.",
        "Expected Output": "Habits are ranked by health impact.\n\nHabits are classified as fragile or resilient.",
        "Actual Output": "Habits are ranked by health impact and correctly classified.",
        "Status": "Pass",
        "More Information": "Checks fragility prediction ML models."
    }
]

# Create a DataFrame
df = pd.DataFrame(test_cases, columns=columns)

# Write to Excel
output_file = "Functional_Test_Cases_Updated.xlsx"

# Using pandas ExcelWriter with openpyxl engine to apply formatting
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Test Cases')
    
    # Auto-adjust column widths for better readability and enable text wrapping
    worksheet = writer.sheets['Test Cases']
    
    # Set header styling
    from openpyxl.styles import Alignment, Font, PatternFill
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for cell in worksheet["1:1"]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Adjust widths and row alignments
    for idx, col in enumerate(df.columns, 1):
        column_letter = get_column_letter(idx)
        # Fixed width for some columns
        if col in ["Feature", "Test Case", "Status"]:
            width = 20
        else:
            width = 45 # wider for description columns
        
        worksheet.column_dimensions[column_letter].width = width
        
        # Enable text wrapping for all cells in the column
        for cell in worksheet[column_letter]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

print(f"Successfully created '{output_file}' with functional test cases.")
