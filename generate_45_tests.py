import pandas as pd
from openpyxl.utils import get_column_letter

# Define the columns based on the provided template
columns = [
    "Feature",
    "Test Case",
    "Steps to execute test case",
    "Expected Output",
    "Actual Output",
    "Status",
    "More Information"
]

# Create exactly 45 granular test cases
raw_test_cases = [
    # User Management - Registration (5)
    ("User Registration", "Navigate to Registration Page", "1. Open application.\n2. Click 'Register'.", "Registration page loads successfully.", "Registration page loads.", "Pass", "Check UI rendering."),
    ("User Registration", "Valid User Registration", "1. Enter valid details.\n2. Click 'Register'.", "Account created successfully.", "Account created.", "Pass", "Check DB."),
    ("User Registration", "Invalid Email Format", "1. Enter email without '@'.\n2. Click 'Register'.", "Validation error displayed.", "Error displayed.", "Pass", "Frontend validation."),
    ("User Registration", "Weak Password Rejection", "1. Enter short password.\n2. Click 'Register'.", "Password strength error shown.", "Error shown.", "Pass", "Backend validation."),
    ("User Registration", "Password Mismatch", "1. Enter non-matching passwords.\n2. Click 'Register'.", "Mismatch error displayed.", "Error displayed.", "Pass", "Frontend validation."),

    # User Management - Login (3)
    ("User Login", "Navigate to Login Page", "1. Open application.\n2. Click 'Login'.", "Login page loads successfully.", "Login page loads.", "Pass", "Check UI rendering."),
    ("User Login", "Valid User Login", "1. Enter valid credentials.\n2. Click 'Login'.", "Redirected to dashboard.", "Redirected to dashboard.", "Pass", "Auth token set."),
    ("User Login", "Invalid Credentials", "1. Enter wrong password.\n2. Click 'Login'.", "Invalid credentials error shown.", "Error shown.", "Pass", "Auth rejection."),

    # Health Profile Setup (7)
    ("Health Profile", "Navigate to Profile Setup", "1. Click 'Profile' in menu.", "Profile setup form loads.", "Form loads.", "Pass", "UI Check."),
    ("Health Profile", "Enter Valid Age", "1. Input age in valid range.", "Age field accepts input.", "Input accepted.", "Pass", "Validation check."),
    ("Health Profile", "Enter Valid Weight", "1. Input weight.", "Weight field accepts input.", "Input accepted.", "Pass", "Validation check."),
    ("Health Profile", "Enter Valid Height", "1. Input height.", "Height field accepts input.", "Input accepted.", "Pass", "Validation check."),
    ("Health Profile", "Auto-calculate BMI", "1. Enter height and weight.", "BMI is automatically calculated.", "BMI calculated correctly.", "Pass", "JS calculation check."),
    ("Health Profile", "Select Activity Level", "1. Choose from dropdown.", "Dropdown selects properly.", "Selection saved.", "Pass", "UI Check."),
    ("Health Profile", "Save Profile", "1. Click 'Save Profile'.", "Profile saved to database.", "Profile saved.", "Pass", "DB verification."),

    # ML Recommendations - Diet (5)
    ("Diet Recommendation", "Trigger Diet Model", "1. Navigate to recommendations.\n2. Click 'Generate Diet'.", "Model starts forming predictions.", "Model triggered.", "Pass", "RF Model init."),
    ("Diet Recommendation", "View BMR Output", "1. Observe BMR reading.", "BMR is displayed based on profile.", "BMR displayed accurately.", "Pass", "Math Verification."),
    ("Diet Recommendation", "View TDEE Output", "1. Observe TDEE reading.", "TDEE is displayed correctly.", "TDEE displayed.", "Pass", "Math Verification."),
    ("Diet Recommendation", "View Calorie Intake", "1. Observe optimal calories.", "Calorie prediction shown.", "Prediction shown.", "Pass", "RF Model output."),
    ("Diet Recommendation", "View Macronutrients", "1. Observe protein/carb/fat split.", "Macros distributed based on goal.", "Macros displayed.", "Pass", "Algorithmic split."),

    # ML Recommendations - Exercise (4)
    ("Exercise Recommendation", "Trigger Exercise Model", "1. Click 'Generate Exercise'.", "Decision tree model triggers.", "Model triggered.", "Pass", "DT Model init."),
    ("Exercise Recommendation", "View Workout Type", "1. Observe workout suggestions.", "Appropriate workouts listed.", "Workouts listed.", "Pass", "Check against fitness level."),
    ("Exercise Recommendation", "View Recommended Duration", "1. Observe duration.", "Duration matches user time.", "Duration matches.", "Pass", "DT logic check."),
    ("Exercise Recommendation", "View Recommended Frequency", "1. Observe weekly frequency.", "Frequency matches fitness goals.", "Frequency matches.", "Pass", "DT logic check."),

    # ML Recommendations - Sleep (4)
    ("Sleep Recommendation", "Trigger Sleep Model", "1. Click 'Generate Sleep'.", "Linear Regression model triggers.", "Model triggered.", "Pass", "LR Model init."),
    ("Sleep Recommendation", "View Predicted Duration", "1. Observe sleep hours.", "Optimal hours predicted.", "Hours predicted.", "Pass", "LR Model output."),
    ("Sleep Recommendation", "View Suggested Bedtime", "1. Observe bedtime.", "Bedtime calculated based on cycles.", "Bedtime shown.", "Pass", "Logic check."),
    ("Sleep Recommendation", "View Wake-up Time", "1. Observe wake-up time.", "Wake-up aligns with bedtime.", "Wake-up shown.", "Pass", "Logic check."),

    # Health Tracking - Daily Log (5)
    ("Health Tracking", "Navigate to Daily Log", "1. Click 'Daily Log'.", "Log entry form loads.", "Form loads.", "Pass", "UI Check."),
    ("Health Tracking", "Enter Daily Weight", "1. Subject weight input.", "Input accepted.", "Input accepted.", "Pass", "Validation."),
    ("Health Tracking", "Enter Sleep Hours", "1. Subject sleep input.", "Input accepted.", "Input accepted.", "Pass", "Validation."),
    ("Health Tracking", "Enter Exercise Minutes", "1. Subject exercise input.", "Input accepted.", "Input accepted.", "Pass", "Validation."),
    ("Health Tracking", "Save Daily Log", "1. Click 'Submit'.", "Data saved successfully.", "Data saved.", "Pass", "DB check."),

    # Analytics Dashboard (2)
    ("Analytics Dashboard", "View Progress Charts", "1. Open dashboard window.", "Charts render with historical data.", "Charts render.", "Pass", "Chart.js check."),
    ("Analytics Dashboard", "View Recent Logs", "1. Scroll to logs table.", "Table shows recent day entries.", "Table populated.", "Pass", "Query check."),

    # Advanced Analytics - Recovery & Stability (4)
    ("Advanced Analytics", "Trigger Recovery Engine", "1. Click 'Recovery Analysis'.", "Engine processes history data.", "Engine processed.", "Pass", "RF Model init."),
    ("Advanced Analytics", "View Recovery Prediction", "1. Observe recovery days result.", "Days predicted properly.", "Days shown.", "Pass", "RF Model output."),
    ("Advanced Analytics", "View Stability Score", "1. Observe stability percentage.", "Score calculates 0-100%.", "Score shown.", "Pass", "GB Classifier check."),
    ("Advanced Analytics", "View Consistency Rate", "1. Observe logging consistency.", "Rate calculates correctly.", "Rate shown.", "Pass", "Metric calculation."),

    # Advanced Analytics - Correlation Engine (3)
    ("Correlation Engine", "Trigger Correlation Analysis", "1. Click 'Correlation Engine'.", "SciPy pearson correlation runs.", "Analysis runs.", "Pass", "Script execution."),
    ("Correlation Engine", "View Top Correlations", "1. Observe behavior matches.", "Highest correlated items split.", "Correlations shown.", "Pass", "Data validity."),
    ("Correlation Engine", "View Root Cause Insights", "1. Observe text insights.", "Insight text maps to parameters.", "Insights generated.", "Pass", "Text generation check."),

    # Advanced Analytics - Habit Sensitivity (3)
    ("Habit Sensitivity", "Trigger Habit Analysis", "1. Click 'Habit Analyzer'.", "Fragility RF Model triggers.", "Model triggers.", "Pass", "RF Model init."),
    ("Habit Sensitivity", "View Fragility Ranking", "1. Observe habit list.", "Habits sorted resilient to fragile.", "Habits sorted.", "Pass", "Output processing."),
    ("Habit Sensitivity", "View High-Impact Alert", "1. Look for high impact tags.", "Most important habits flagged.", "Habits flagged.", "Pass", "Analysis check.")
]

test_cases = []
for item in raw_test_cases:
    test_cases.append({
        "Feature": item[0],
        "Test Case": item[1],
        "Steps to execute test case": item[2],
        "Expected Output": item[3],
        "Actual Output": item[4],
        "Status": item[5],
        "More Information": item[6]
    })

# Verify exactly 45 items
if len(test_cases) != 45:
    print(f"Warning: Count is {len(test_cases)}, not 45.")

# Create a DataFrame
df = pd.DataFrame(test_cases, columns=columns)

# Write to Excel
output_file = "Functional_Test_Cases_Expanded.xlsx"

# Using pandas ExcelWriter with openpyxl engine to apply formatting
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Test Cases')
    
    # Auto-adjust column widths for better readability and enable text wrapping
    worksheet = writer.sheets['Test Cases']
    
    # Set header styling
    from openpyxl.styles import Alignment, Font, PatternFill
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for cell in worksheet["1:1"]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Adjust widths and row alignments
    for idx, col in enumerate(df.columns, 1):
        column_letter = get_column_letter(idx)
        if col in ["Feature", "Test Case", "Status"]:
            width = 25
        else:
            width = 40 
        
        worksheet.column_dimensions[column_letter].width = width
        
        for cell in worksheet[column_letter]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

print(f"Successfully created '{output_file}' with {len(test_cases)} functional test cases.")
